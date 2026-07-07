#!/usr/bin/env python3
"""
parse_xlsx.py  —  Doomlings card-set build tool (Excel workbook edition).

Replaces the old CSV parser. Reads a workbook with three worksheets:
  * Traits        — Name, Category, Color, Point Face Value, At World's End,
                    Action Card, Drop of Life, Dominant Trait, Rarity, Effect
  * Ages          — Name, Category, Rarity, Effect
  * Catastrophes  — Name, Category, Rarity, Stage 1, Stage 2, Final Catastrophe

Produces data/<set-id>.json (data only, no logic). Each trait's scoring text is
turned into STRUCTURED, REUSABLE effect objects (see effects.js). Anything it
can't confidently parse lands in `needsReview` instead of being guessed.

Run:  python tools/parse_xlsx.py [workbook.xlsx] [set-id] "Set Name"
      (defaults to doomlings_base_game_cards.xlsx / base-game / "Base Game")
"""

import json
import re
import sys
from pathlib import Path

import openpyxl

COLOR_MAP = {
    "blue": "blue", "green": "green", "red": "red", "purple": "purple",
    "colorless (gray)": "colorless", "colorless": "colorless", "": "", None: "",
}
COLOR_WORDS = "(blue|green|red|purple|colorless)"


def norm_color(raw):
    return COLOR_MAP.get((raw or "").strip().lower(), (raw or "").strip().lower())


def yn(v):
    return str(v or "").strip().lower() == "yes"


def txt(v):
    return "" if v is None else str(v).strip()


# --------------------------------------------------------------------------
# Trait scoring parse -> (effects, choice, needsReview)
# --------------------------------------------------------------------------
def parse_scoring_effects(name, effect, atWorldsEnd, dropOfLife):
    text = txt(effect)
    low = text.lower()
    effects = []
    matched = [False]

    def add(e):
        effects.append(dict(e, text=text))
        matched[0] = True

    # ---- At World's End choices (resolved interactively; no static score) ----
    choice = None
    if atWorldsEnd:
        cmap = {
            "FAITH": "faith",
            "HYPER-INTELLIGENCE": "hyperIntelligence",
            "PREPPER": "prepper",
            "CHERISHED": "cherished",
            "SENTIENCE": "sentience",   # choose a color -> +1 per own trait of it
            "VIRAL": "viral",           # choose a color -> opponents -1 per it
        }
        if name in cmap:
            choice = {"type": cmap[name]}
        # ELOQUENCE / SNEAKY: no structured choice -> surfaced as a manual note.

    # ---- "Value is equal to …" cards (face value replaced by a computed sum) --
    if "value is equal to" in low:
        if "gene pool" in low:
            add({"kind": "valueEqualsGenePool"})
        elif "cards in your hand with effects" in low:
            add({"kind": "valueEqualsHandEffectCards"})
        elif "cards in your hand" in low:
            add({"kind": "valueEqualsHandCount"})
        elif "different colors in your trait pile" in low or "colors in your trait pile" in low:
            add({"kind": "valueEqualsDistinctPileColors"})

    # ---- additive drop-of-life patterns -------------------------------------
    # "+N for every M <color> traits"
    m = re.search(r"([+-]?\d+)\s+for every\s+(\d+)\s+" + COLOR_WORDS + r"\s+traits", low)
    if m:
        add({"kind": "perNColorTrait", "color": m.group(3), "n": int(m.group(2)), "amount": int(m.group(1))})

    # "+N for every pair of <color> traits in each opponent's trait pile" (Branches)
    m = re.search(r"([+-]?\d+)\s+for every pair of\s+" + COLOR_WORDS + r"\s+traits in each opponent", low)
    if m:
        add({"kind": "perOpponentColorPair", "color": m.group(2), "n": 2, "amount": int(m.group(1))})

    # "+N for every color pair in your trait pile" (Pack Behavior)
    m = re.search(r"([+-]?\d+)\s+for every color pair", low)
    if m:
        add({"kind": "perSameColorGroup", "n": 2, "amount": int(m.group(1))})

    # "+N for every M traits of the same color"
    m = re.search(r"([+-]?\d+)\s+for every\s+(\d+)\s+traits of the same color", low)
    if m:
        add({"kind": "perSameColorGroup", "n": int(m.group(2)), "amount": int(m.group(1))})

    # "+N for each <color> trait"
    m = re.search(r"([+-]?\d+)\s+for each\s+" + COLOR_WORDS + r"\s+trait", low)
    if m:
        add({"kind": "perColorTrait", "color": m.group(2), "amount": int(m.group(1))})

    # "+N for each negative face value trait" (Immunity)
    m = re.search(r"([+-]?\d+)\s+for each negative face value trait", low)
    if m:
        add({"kind": "perNegativeTrait", "amount": int(m.group(1))})

    # "+N for each face value (V) trait" (Pollination)
    m = re.search(r"([+-]?\d+)\s+for each face value\s*\((\d+)\)\s+trait", low)
    if m:
        add({"kind": "perFaceValueTrait", "value": int(m.group(2)), "amount": int(m.group(1))})

    # "-N for each trait in your trait pile" (Tiny)
    m = re.search(r"([+-]?\d+)\s+for each trait in your trait pile", low)
    if m:
        add({"kind": "perTraitInPile", "amount": int(m.group(1)), "excludeDominants": "excluding dominant" in low})

    # "+N for each trait of the lowest color count" (Symbiosis)
    m = re.search(r"([+-]?\d+)\s+for each trait of the lowest color count", low)
    if m:
        add({"kind": "perRarestColorTrait", "amount": int(m.group(1)), "requiresMultiColor": True})

    # "+N if you have the most traits in your trait pile" (Apex Predator)
    m = re.search(r"([+-]?\d+)\s+if you have the most traits", low)
    if m:
        add({"kind": "flatIfMostTraits", "amount": int(m.group(1))})

    # "+N for each <Named> in all trait piles" (Swarm)
    m = re.search(r"([+-]?\d+)\s+for each\s+(\w+)\s+in all trait piles", low)
    if m:
        add({"kind": "perNamedTrait", "name": m.group(2).lower(), "scope": "all", "amount": int(m.group(1))})

    # "+N for each <Named> in your trait pile" (Kidney) — only if not already a color pattern
    m = re.search(r"([+-]?\d+)\s+for each\s+(\w+)\s+in your trait pile", low)
    if m and not matched[0]:
        add({"kind": "perNamedTrait", "name": m.group(2).lower(), "scope": "self", "amount": int(m.group(1))})

    # hand-based
    m = re.search(r"([+-]?\d+)\s+for each dominant card in your hand", low)   # Brave
    if m:
        add({"kind": "perDominantInHand", "amount": int(m.group(1))})
    m = re.search(r"([+-]?\d+)\s+for each (?:different )?color in your hand", low)  # Saudade
    if m:
        add({"kind": "perColorInHand", "amount": int(m.group(1))})
    m = re.search(r"([+-]?\d+)\s+for each card in your hand", low)            # Camouflage/Fortunate
    if m:
        add({"kind": "perCardInHand", "amount": int(m.group(1))})

    # --- needs-review detection ---------------------------------------------
    scoreish = dropOfLife or (atWorldsEnd and choice is None)
    scoring_words = re.search(
        r"[+-]\s*\d|gene pool|for each|for every|most traits|lowest color|"
        r"missing|value is equal|color pair|for all traits", low)
    review = bool(scoreish and scoring_words and not matched[0] and choice is None)
    return effects, choice, review


# --------------------------------------------------------------------------
# Catastrophe final effect -> structured World's End effect
# --------------------------------------------------------------------------
def parse_catastrophe_final(final_text):
    low = txt(final_text).lower()
    t = txt(final_text)

    # "-N for each <color> trait …" (Ice Age, Retrovirus, Solar Flare, Super Volcano)
    m = re.search(r"([+-]?\d+)\s+for each\s+" + COLOR_WORDS + r"\s+trait", low)
    if m:
        return {"kind": "perColorTraitPoints", "color": m.group(2), "amount": int(m.group(1)), "text": t}

    # "-N … for each color missing from your trait pile" (The Big One)
    m = re.search(r"([+-]?\d+)\s+(?:to your score\s+)?for each color missing", low)
    if m:
        return {"kind": "perMissingColorPoints", "amount": int(m.group(1)), "text": t}

    # "+N points to the player(s) with the fewest traits" (Overpopulation)
    m = re.search(r"([+-]?\d+)\s+points to the player\(s\) with the fewest traits", low)
    if m:
        return {"kind": "pointsToFewestTraits", "amount": int(m.group(1)), "text": t}

    # "-N points to the player(s) with the most traits" (Grey Goo)
    m = re.search(r"([+-]?\d+)\s+points to the player\(s\) with the most traits", low)
    if m:
        return {"kind": "pointsToMostTraits", "amount": int(m.group(1)), "text": t}

    # Deus Ex Machina — draw a card, add face value, capped
    if "add its face value to your final score" in low:
        m = re.search(r"(\d+)\s*max", low) or re.search(r"max\s*\+?(\d+)", low)
        return {"kind": "deusExMachina", "max": int(m.group(1)) if m else 7, "text": t}

    # AI Takeover — colorless traits now worth 2, colorless effects ignored
    if "colorless" in low and ("now worth 2" in low or "colorless traits to 2" in low):
        return {"kind": "aiTakeover", "text": t}

    # "Discard 1 <color> trait from your trait pile"
    m = re.search(r"discard 1\s+" + COLOR_WORDS + r"\s+trait from your trait pile", low)
    if m:
        return {"kind": "discardColorTrait", "color": m.group(1), "text": t}

    # "Discard 1 trait … with a face value of N or higher" (Four Horsemen)
    m = re.search(r"face value\s+(?:of\s+)?\(?(\d+)\)?\s*(?:trait\s+)?or higher", low)
    if m:
        return {"kind": "discardHighFaceTrait", "minValue": int(m.group(1)), "text": t}

    return {"kind": "unparsed", "text": t}


# --------------------------------------------------------------------------
def build(path, set_id, set_name):
    wb = openpyxl.load_workbook(path, data_only=True)
    traits, ages, catastrophes, needs_review = [], [], [], []

    # --- Traits (data rows start at 3: title, header, then cards) ---
    for r in wb["Traits"].iter_rows(min_row=3, values_only=True):
        name = txt(r[0])
        if not name:
            continue
        face_raw = r[3]
        face = int(face_raw) if isinstance(face_raw, (int, float)) else 0
        atWE, dol = yn(r[4]), yn(r[6])
        card = {
            "name": name,
            "color": norm_color(r[2]),
            "faceValue": face,
            "atWorldsEnd": atWE,
            "isAction": yn(r[5]),
            "dropOfLife": dol,
            "isDominant": yn(r[7]),
            "rarity": txt(r[8]),
            "effect": txt(r[9]),
        }
        eff, choice, review = parse_scoring_effects(name, r[9], atWE, dol)
        card["scoringEffects"] = eff
        if choice:
            card["worldsEndChoice"] = choice
        traits.append(card)
        if review:
            needs_review.append({"name": name, "effect": card["effect"]})

    # --- Ages (Name, Category, Rarity, Effect) ---
    for r in wb["Ages"].iter_rows(min_row=3, values_only=True):
        name = txt(r[0])
        if not name:
            continue
        ages.append({"name": name, "rarity": txt(r[2]), "effect": txt(r[3])})

    # --- Catastrophes (Name, Category, Rarity, Stage1, Stage2, Final) ---
    for r in wb["Catastrophes"].iter_rows(min_row=2, values_only=True):
        name = txt(r[0])
        if not name or txt(r[1]).lower() != "catastrophe":
            continue
        catastrophes.append({
            "name": name,
            "color": "",
            "rarity": txt(r[2]),
            "genePoolEffect": txt(r[3]),
            "stage2Effect": txt(r[4]),
            "finalEffect": txt(r[5]),
            "worldsEnd": parse_catastrophe_final(r[5]),
        })

    return {
        "setId": set_id,
        "setName": set_name,
        "colors": ["blue", "green", "red", "purple", "colorless"],
        "traits": traits,
        "ages": ages,
        "catastrophes": catastrophes,
        "needsReview": needs_review,
    }


if __name__ == "__main__":
    here = Path(__file__).resolve().parent
    default = here.parent / "doomlings_base_game_cards.xlsx"
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else default
    set_id = sys.argv[2] if len(sys.argv) > 2 else "base-game"
    set_name = sys.argv[3] if len(sys.argv) > 3 else "Base Game"

    data = build(path, set_id, set_name)
    out = here.parent / "data" / f"{set_id}.json"
    out.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"Wrote {out}")
    print(f"  traits: {len(data['traits'])}  ages: {len(data['ages'])}  "
          f"catastrophes: {len(data['catastrophes'])}  needs-review: {len(data['needsReview'])}")
    for r in data["needsReview"]:
        print(f"    review: {r['name']}: {r['effect']}")
    for c in data["catastrophes"]:
        if c["worldsEnd"]["kind"] == "unparsed":
            print(f"    [unparsed catastrophe] {c['name']}: {c['finalEffect']}")
